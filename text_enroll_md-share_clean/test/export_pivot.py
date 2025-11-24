#!/usr/bin/env python3

import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import re
import numpy as np

result_root_dir = sys.argv[1]
output_xlsx = sys.argv[2]

testdata_order = [
    'magic_7kw_aishell4',
    'fyt_7zh_16spk_aishell4',
    'fyt_48zh_yue_aishell4',
    'fyt_48zh_yue_24zh_w_20zh_aishell4',
    'fyt_48zh_hebei_aishell4',
    'fyt_48zh_hebei_24zh_w_20zh_aishell4',
    'fyt_20zh_qingdao_aishell4',
    'fyt_22zh_henan_aishell4',
    'fyt_15zh_chongqing_aishell4',
    'fyt_20zh_context_aishell4'
]

test_option_order = ['1_1_0_0_1', '0_1_0_0_0']

# result selection for an experiment

# accent dialect finetune
target_subdir_pattern = re.compile(r"5000h_tang_ke_sinvxx_1-3s_lex_ft_.*|5000h_tang_ke_sinvxx_1-3s_lex_hanlp_.*|.*hebei.*|.*qingdao.*|.*henan.*|.*chongqing.*")

train_data_pattern = r'(.*_hanlp)'
model_arch_pattern = r'(transformer_.*)'
train_process_pattern = r'(reverb_perturb_.*)'
epoch_pattern = r'(avg[0-9]*_[0-9]*)'
testdata_pattern = r'(.*_aishell4)'

subdir_pattern = re.compile(r"{}_{}_{}_{}-{}".format(train_data_pattern, model_arch_pattern, train_process_pattern, epoch_pattern, testdata_pattern))

test_option_pattern = r'((?<=_)[01]_[01]_[01]_[01]_[01](?=_))'


def load_df_row(input_csv):
    sub_dir = os.path.dirname(input_csv)
    sub_dir = os.path.basename(sub_dir)
    file_name = os.path.basename(input_csv)
    test_option = re.findall(test_option_pattern, file_name)[0]
    match = subdir_pattern.match(sub_dir)
    if match:
        train_data = match.group(1)
        model_arch = match.group(2)
        train_process = match.group(3)
        epoch = match.group(4)
        testdata = match.group(5)
        model = "{}_{}_{}_{}".format(train_data, model_arch, train_process, epoch)
    else:
        print("Error: subdir pattern not match. sub_dir: {}".format(sub_dir))
        sys.exit(1)

    df = pd.read_csv(input_csv)
    assert list(df.columns) == ['keyword', 'auc', 'cost_miss', 'cost_fa', 'prior', 'dcf', 'best_thres', 'p_miss', 'p_fa', 'fa_per_hour']
    rows = df[df['keyword'] == 'avg']
    assert len(rows) == 1
    avg_row = rows.iloc[0]

    df_row = {
        'model': model,
        'train_data': train_data,
        'model_arch': model_arch,
        'train_process': train_process,
        'epoch': epoch,
        'testdata': testdata,
        'test_option': test_option,
        'cost_miss': avg_row['cost_miss'],
        'cost_fa': avg_row['cost_fa'],
        'prior': avg_row['prior'],
        'auc_of_roc': avg_row['auc'],
        'dcf': avg_row['dcf'],
        'threshold': avg_row['best_thres'],
        'p_miss': avg_row['p_miss'],
        'p_fa': avg_row['p_fa'],
        'fa_per_h': avg_row['fa_per_hour']
    }
    return df_row


def reformat_xlsx(output_xlsx):
    # 加载Excel文件
    wb = load_workbook(output_xlsx)
    ws = wb.active

    # 创建样式（保留n位小数、百分比、科学记数法）
    style_decimal4 = NamedStyle(name="decimal4", number_format="0.0000")
    style_decimal1 = NamedStyle(name="decimal1", number_format="0.0")
    style_percent = NamedStyle(name="percent", number_format="0.00%")
    style_scientific = NamedStyle(name="scientific", number_format="0.000E+00")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        columns = list(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column))
        column_names = [cell.value for cell in columns[0]]
        print(column_names)
        for col_idx, col_name in enumerate(column_names, 1):
            if col_name == None:
                continue
            elif 'auc_of_roc' in col_name:
                style = style_decimal4
            elif 'dcf' in col_name:
                style = style_scientific
            elif 'threshold' in col_name:
                style = style_decimal4
            elif 'p_miss' in col_name:
                style = style_percent
            elif 'fa_per_h' in col_name:
                style = style_decimal1
            else:
                style = style_decimal4
                continue
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.style = style

    # 保存Excel文件
    wb.save(output_xlsx)


def main():
    df = pd.DataFrame()
    for root, dirs, files in os.walk(result_root_dir):
        if root != result_root_dir:
            continue
        for dir in dirs:
            if target_subdir_pattern.match(dir) == None:
                continue
            for root, dirs, files in os.walk(os.path.join(result_root_dir, dir)):
                if root != os.path.join(result_root_dir, dir):
                    continue
                for file in files:
                    if re.findall("dcf_auc.*_avg.csv", file) == []:
                        continue
                    input_csv = os.path.join(result_root_dir, dir, file)
                    row = load_df_row(input_csv)

                    df = pd.concat([df, pd.DataFrame([row])])

    index = ['cost_miss', 'cost_fa', 'prior', 'test_option', 'train_process', 'model_arch', 'train_data', 'epoch', 'model', 'testdata']
    values = ['auc_of_roc', 'dcf', 'threshold', 'p_miss', 'fa_per_h']

    df['testdata'] = pd.Categorical(df['testdata'], categories=testdata_order, ordered=True)
    df['test_option'] = pd.Categorical(df['test_option'], categories=test_option_order, ordered=True)
    df['train_data'] = pd.Categorical(df['train_data'], ordered=True)
    df['model_arch'] = pd.Categorical(df['model_arch'], ordered=True)
    df['train_process'] = pd.Categorical(df['train_process'], ordered=True)

    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        for cost_miss in df['cost_miss'].unique():
            for cost_fa in df['cost_fa'].unique():
                for piror in df['prior'].unique():
                    for test_option in df['test_option'].unique():

                        # 生成透视表
                        pivot = df.pivot_table(
                            index=index,
                            values=values,
                            aggfunc=lambda x: np.nan if len(x) == 0 else np.sum(x)
                        )
                        pivot = pivot[values]
                        pivot = pivot.loc[pivot.index.get_level_values('cost_miss') == cost_miss]
                        pivot = pivot.loc[pivot.index.get_level_values('cost_fa') == cost_fa]
                        pivot = pivot.loc[pivot.index.get_level_values('prior') == piror]
                        pivot = pivot.loc[pivot.index.get_level_values('test_option') == test_option]
                        pivot = pivot.reset_index()
                        pivot = pivot.drop(columns='cost_miss')
                        pivot = pivot.drop(columns='cost_fa')
                        pivot = pivot.drop(columns='prior')
                        pivot = pivot.drop(columns='test_option')
                        pivot.set_index(['train_process', 'model_arch', 'train_data', 'epoch', 'model', 'testdata'], inplace=True)
                        # print(pivot)
                        
                        # 写入Excel
                        sheet_name = "{}_{}_{}_{}".format(cost_miss, cost_fa, piror, test_option)[:31]  # Excel sheet名称长度限制
                        print(sheet_name)
                        pivot.to_excel(writer, sheet_name=sheet_name)

    reformat_xlsx(output_xlsx)


if __name__ == "__main__":
    main()
